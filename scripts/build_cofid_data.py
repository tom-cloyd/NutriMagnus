#!/usr/bin/env python3
"""
build_cofid_data.py — regenerate cofid_data.json from the published UK
McCance and Widdowson's Composition of Foods Integrated Dataset (CoFID) 2021
spreadsheet.

CoFID is a static download, not a live API (see cofid_lookup.py's module
docstring for why that makes it a "static-dataset" source rather than a
"live" one like USDA/OFF/CNF) — this script is the one-time ingest step, run
by a developer, not by numa at runtime.

Usage:
    pip install openpyxl   # not a runtime dependency — only this script needs it
    python scripts/build_cofid_data.py path/to/CoFID_2021.xlsx

Source spreadsheet: "McCance and Widdowson's composition of foods integrated
dataset" (2021), published at
https://www.gov.uk/government/publications/composition-of-foods-integrated-dataset-cofid
Not redistributed here — download it yourself and pass the path in.

CoFID's own data is spread across 14 sheets ("1.3 Proximates", "1.4
Inorganics", "1.5 Vitamins", ... fatty-acid breakdowns, phytosterols, organic
acids). Only the three merged below map cleanly onto numa's shared nutrient
vocabulary. Notably: CoFID has NO amino-acid sheet — unlike CNF, it
contributes nothing to the aa_* keys DIAAS calculations depend on. Values are
per 100 g of food (not per 100g fatty acid, where a sheet offers both).

Non-numeric cells use McCance & Widdowson's standard annotations: "Tr"
(trace, not quantified) and "N" (no reliable data). Both are simply skipped
(float() raises, caught below) rather than guessed at.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("This script needs openpyxl: pip install openpyxl (not a runtime dependency)")
    sys.exit(1)

OUTPUT = Path(__file__).parent.parent / "cofid_data.json"

# (sheet name, header row with column labels, {column label -> (our_key, multiplier)})
# Multiplier converts CoFID's unit to ours; all of these already match (g/mg/µg).
_SHEETS: list[tuple[str, dict[str, tuple[str, float]]]] = [
    ("1.3 Proximates", {
        "Protein (g)":             ("protein_g",        1.0),
        "Fat (g)":                 ("fat_g",             1.0),
        "Carbohydrate (g)":        ("carbs_g",           1.0),
        "Energy (kcal) (kcal)":    ("calories",          1.0),
        "AOAC fibre (g)":          ("fiber_g",           1.0),
        "Total sugars (g)":        ("sugar_g",           1.0),
        "Satd FA /100g fd (g)":    ("saturated_fat_g",   1.0),
        "Mono FA /100g food (g)":  ("mono_fat_g",         1.0),
        "Poly FA /100g food (g)":  ("poly_fat_g",         1.0),
    }),
    ("1.4 Inorganics", {
        "Sodium (mg)":     ("sodium_mg",     1.0),
        "Potassium (mg)":  ("potassium_mg",  1.0),
        "Calcium (mg)":    ("calcium_mg",    1.0),
        "Magnesium (mg)":  ("magnesium_mg",  1.0),
        "Phosphorus (mg)": ("phosphorus_mg", 1.0),
        "Iron (mg)":       ("iron_mg",       1.0),
        "Zinc (mg)":       ("zinc_mg",       1.0),
        "Selenium (µg)":   ("selenium_mcg",  1.0),
        "Iodine (µg)":     ("iodine_mcg",    1.0),
    }),
    ("1.5 Vitamins", {
        # Retinol Equivalent, unlike CNF's retinol-only figure, already
        # accounts for provitamin-A carotenoids — the more correct vitamin_a_mcg.
        "Retinol Equivalent (µg)": ("vitamin_a_mcg", 1.0),
        "Carotene (µg)":           ("beta_carotene_mcg", 1.0),  # total carotene, not beta- specifically
        "Vitamin D (µg)":          ("vitamin_d_mcg", 1.0),
        "Vitamin E (mg)":          ("vitamin_e_mg",  1.0),
        "Vitamin K1 (µg)":         ("vitamin_k_mcg", 1.0),
        "Thiamin (mg)":            ("thiamin_mg",    1.0),
        "Riboflavin (mg)":         ("riboflavin_mg", 1.0),
        "Niacin (mg)":             ("niacin_mg",     1.0),
        "Vitamin B6 (mg)":         ("b6_mg",         1.0),
        "Vitamin B12 (µg)":        ("b12_mcg",       1.0),
        "Folate (µg)":             ("folate_mcg",    1.0),
        "Vitamin C (mg)":          ("vitamin_c_mg",  1.0),
    }),
]


def _parse_sheet(ws, column_map: dict[str, tuple[str, float]]) -> dict[str, dict]:
    """Row 1 has header labels, row 4+ has data (rows 2-3 are internal codes/
    long names we don't need). Returns {food_code: {our_key: value}}."""
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {label: i for i, label in enumerate(header) if label in column_map}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        food_code = row[0]
        if not food_code:
            continue
        nutrients: dict[str, float] = {}
        for label, i in col_index.items():
            our_key, multiplier = column_map[label]
            raw = row[i] if i < len(row) else None
            if raw is None:
                continue
            try:
                nutrients[our_key] = float(raw) * multiplier
            except (TypeError, ValueError):
                continue  # "Tr", "N", or otherwise non-numeric — skip, don't guess
        out[food_code] = {"name": row[1], "nutrients": nutrients}
    return out


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    merged: dict[str, dict] = {}
    for sheet_name, column_map in _SHEETS:
        parsed = _parse_sheet(wb[sheet_name], column_map)
        for food_code, entry in parsed.items():
            if food_code not in merged:
                merged[food_code] = {"food_code": food_code, "name": entry["name"], "nutrients": {}}
            merged[food_code]["nutrients"].update(entry["nutrients"])

    records = sorted(merged.values(), key=lambda r: r["food_code"])
    OUTPUT.write_text(json.dumps(records, indent=1), encoding="utf-8")
    print(f"Wrote {len(records)} foods to {OUTPUT} ({OUTPUT.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
