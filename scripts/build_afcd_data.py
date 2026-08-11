#!/usr/bin/env python3
"""
build_afcd_data.py — regenerate afcd_data.json from the published Australian
Food Composition Database (AFCD) Release 3 "Nutrient profiles" spreadsheet.

AFCD is a static download, not a live API — this script is the one-time
ingest step, run by a developer, not by numa at runtime (see
afcd_lookup.py's module docstring for the static-vs-live source distinction).

Usage:
    pip install openpyxl   # not a runtime dependency — only this script needs it
    python scripts/build_afcd_data.py "path/to/AFCD Release 3 - Nutrient profiles.xlsx"

Source spreadsheet: Food Standards Australia New Zealand (FSANZ), "AFCD
Release 3 - Nutrient profiles", published at
https://www.foodstandards.gov.au/science-data/food-nutrient-databases/afcd/data-files
Not redistributed here — download it yourself and pass the path in. Uses the
"All solids & liquids per 100 g" sheet only (there's a separate "per 100 mL"
sheet for liquids that this doesn't touch).

Unlike CoFID, AFCD DOES publish amino acids (in mg per 100g, converted to g
here to match numa's aa_* keys) — the manual's "notably deep on amino acid
profiles" claim for AFCD holds up. Cystine and cysteine are reported as one
combined figure ("Cystine plus cysteine"), mapped to aa_cystine_g as the
closest approximation. Energy is published in kJ only; converted to kcal
here via the standard 4.184 kJ/kcal factor.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("This script needs openpyxl: pip install openpyxl (not a runtime dependency)")
    sys.exit(1)

OUTPUT = Path(__file__).parent.parent / "afcd_data.json"
KJ_PER_KCAL = 4.184

# {column label -> (our_key, multiplier)}. Multiplier converts AFCD's unit to
# ours — mostly 1.0; amino acids are mg in AFCD, g in our schema (÷1000).
_COLUMN_MAP: dict[str, tuple[str, float]] = {
    "Protein \n(g)":                                          ("protein_g",        1.0),
    "Fat, total \n(g)":                                       ("fat_g",            1.0),
    "Available carbohydrate, without sugar alcohols \n(g)":   ("carbs_g",          1.0),
    "Total dietary fibre \n(g)":                              ("fiber_g",          1.0),
    "Total sugars (g)":                                       ("sugar_g",          1.0),
    "Total saturated fatty acids, equated \n(g)":             ("saturated_fat_g",  1.0),
    "Total monounsaturated fatty acids, equated \n(g)":       ("mono_fat_g",       1.0),
    "Total polyunsaturated fatty acids, equated \n(g)":       ("poly_fat_g",       1.0),
    "C18:3w3 (g)":                                            ("omega3_ala_mg",   1000.0),
    "C20:5w3 (mg)":                                           ("omega3_epa_mg",   1.0),
    "C22:6w3 (mg)":                                           ("omega3_dha_mg",   1.0),
    "C18:2w6 (g)":                                            ("omega6_la_mg",    1000.0),
    "Calcium (Ca) \n(mg)":                                    ("calcium_mg",       1.0),
    "Iron (Fe) \n(mg)":                                       ("iron_mg",          1.0),
    "Magnesium (Mg) \n(mg)":                                  ("magnesium_mg",     1.0),
    "Phosphorus (P) \n(mg)":                                  ("phosphorus_mg",    1.0),
    "Potassium (K) \n(mg)":                                   ("potassium_mg",     1.0),
    "Sodium (Na) \n(mg)":                                     ("sodium_mg",        1.0),
    "Zinc (Zn) \n(mg)":                                       ("zinc_mg",          1.0),
    "Iodine (I) \n(ug)":                                      ("iodine_mcg",       1.0),
    "Selenium (Se) \n(ug)":                                   ("selenium_mcg",     1.0),
    # Retinol equivalents, unlike a retinol-only figure, already accounts
    # for provitamin-A carotenoids.
    "Vitamin A retinol equivalents \n(ug)":                   ("vitamin_a_mcg",    1.0),
    "Alpha-carotene \n(ug)":                                  ("alpha_carotene_mcg", 1.0),
    "Beta-carotene \n(ug)":                                   ("beta_carotene_mcg", 1.0),
    "Lycopene \n(ug)":                                        ("lycopene_mcg",     1.0),
    "Lutein \n(ug)":                                          ("lutein_zeaxanthin_mcg", 1.0),  # lutein only, no zeaxanthin figure
    "Vitamin C \n(mg)":                                       ("vitamin_c_mg",     1.0),
    "Vitamin D3 equivalents \n(ug)":                          ("vitamin_d_mcg",    1.0),
    "Vitamin E \n(mg)":                                       ("vitamin_e_mg",     1.0),
    "Thiamin (B1) \n(mg)":                                    ("thiamin_mg",       1.0),
    "Riboflavin (B2) \n(mg)":                                 ("riboflavin_mg",    1.0),
    "Niacin (B3) \n(mg)":                                     ("niacin_mg",        1.0),
    "Pyridoxine (B6) \n(mg)":                                 ("b6_mg",            1.0),
    "Cobalamin (B12) \n(ug)":                                 ("b12_mcg",          1.0),
    "Dietary folate equivalents \n(ug)":                      ("folate_mcg",       1.0),
    "Histidine \n(mg)":                                       ("aa_histidine_g",     0.001),
    "Isoleucine \n(mg)":                                      ("aa_isoleucine_g",    0.001),
    "Leucine \n(mg)":                                         ("aa_leucine_g",       0.001),
    "Lysine \n(mg)":                                          ("aa_lysine_g",        0.001),
    "Methionine \n(mg)":                                      ("aa_methionine_g",    0.001),
    "Phenylalanine \n(mg)":                                   ("aa_phenylalanine_g", 0.001),
    "Threonine \n(mg)":                                       ("aa_threonine_g",     0.001),
    "Tyrosine \n(mg)":                                        ("aa_tyrosine_g",      0.001),
    "Tryptophan \n(mg)":                                      ("aa_tryptophan_g",    0.001),
    "Valine \n(mg)":                                          ("aa_valine_g",        0.001),
    "Cystine plus cysteine \n(mg)":                           ("aa_cystine_g",       0.001),  # combined figure
}
_ENERGY_KJ_COLUMN = "Energy with dietary fibre, equated \n(kJ)"


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["All solids & liquids per 100 g"]
    header = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    col_index = {label: i for i, label in enumerate(header) if label in _COLUMN_MAP}
    energy_i = header.index(_ENERGY_KJ_COLUMN)

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        food_code = row[0]
        if not food_code:
            continue
        nutrients: dict[str, float] = {}
        kj = row[energy_i]
        if kj is not None:
            try:
                nutrients["calories"] = float(kj) / KJ_PER_KCAL
            except (TypeError, ValueError):
                pass
        for label, i in col_index.items():
            our_key, multiplier = _COLUMN_MAP[label]
            raw = row[i] if i < len(row) else None
            if raw is None:
                continue
            try:
                nutrients[our_key] = float(raw) * multiplier
            except (TypeError, ValueError):
                continue  # non-numeric annotation — skip, don't guess
        records.append({"food_code": food_code, "name": row[3], "nutrients": nutrients})

    records.sort(key=lambda r: r["food_code"])
    OUTPUT.write_text(json.dumps(records, indent=1), encoding="utf-8")
    print(f"Wrote {len(records)} foods to {OUTPUT} ({OUTPUT.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
